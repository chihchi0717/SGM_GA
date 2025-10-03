import os
import pandas as pd
import numpy as np
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import sys

sys.stdout.reconfigure(encoding="utf-8")

# === 主要設定：只需要修改這裡 ===
# 設定包含 EXP、SIM_design 和 SIM_shrinkage 資料夾的上層路徑
base_folder = r"C:\Users\cchih\Desktop\NTHU\MasterThesis\research_log\best_params\DOE\design[0.6, 0.6, 60]"

# === 自動生成路徑 ===
exp_folder = os.path.join(base_folder, "EXP")
design_sim_folder = os.path.join(
    base_folder, "SIM_design"
)  # "SIM_design" #blank #SIM_shrinkage
shrinkage_sim_folder = os.path.join(base_folder, "SIM_fillet")
output_path = os.path.join(
    base_folder, "EXP_SIM_design.xlsx"
)  # "EXP_SIM_combined.xlsx"

# === 檢查路徑是否存在 ===
if not os.path.isdir(exp_folder):
    print(f"❌ 錯誤：實驗資料夾不存在於 {exp_folder}")
    sys.exit()  # 如果主要資料夾不存在，則終止程式


# === 仰角範圍 ===
angles = list(range(10, 90, 10))

# === 建立 Excel 檔案 ===
wb = Workbook()
wb.remove(wb.active)  # 移除預設空白工作表

for angle in angles:
    angle_str = str(angle)
    angle_tag = f"ele{angle}"  # EXP檔名關鍵字
    polar_tag = f"polar-{angle}.txt"  # SIM檔案名稱

    # === 收集該仰角的所有實驗檔案 ===
    angle_exp_files = sorted([f for f in os.listdir(exp_folder) if angle_tag in f])

    # === 依照試片編號（_01_, _02_）分群 ===
    pattern = re.compile(r"_0*(\d+S?)_BLP")
    sample_groups = defaultdict(list)
    for f in angle_exp_files:
        match = pattern.search(f)
        if match:
            # 直接使用字串作為 ID，例如 '1' 或 '1S'
            sample_id = match.group(1)
            sample_groups[sample_id].append(f)

    # === 讀取實驗資料 ===
    exp_values = {}
    degree_column = None

    for sample_id in sorted(sample_groups):
        files = sorted(sample_groups[sample_id])  # 按時間順序
        for idx, file in enumerate(files):
            df = pd.read_excel(os.path.join(exp_folder, file))
            col_name = df.columns[0]
            intensity_col = df.columns[1]
            if degree_column is None:
                degree_column = df[col_name].values
            exp_name = f"EXP{sample_id}-{idx+1}"
            exp_values[exp_name] = df[intensity_col].values

    # === 建立工作表資料框（包含 EXP_angle）===
    data = {}
    if degree_column is not None:
        data["EXP_angle"] = degree_column
    else:
        data["EXP_angle"] = []

    data.update(exp_values)

    # === 讀取 Design 模擬資料 ===
    sim_angle_column = None  # 存模擬角度
    if os.path.exists(design_sim_folder):
        design_sim_structures = sorted(
            [
                d
                for d in os.listdir(design_sim_folder)
                if os.path.isdir(os.path.join(design_sim_folder, d))
            ]
        )
        for sim_dir in design_sim_structures:
            polar_path = os.path.join(design_sim_folder, sim_dir, polar_tag)
            if os.path.exists(polar_path):
                degs, intensities = [], []
                with open(polar_path, "r") as f:
                    for line in f:
                        parts = line.strip().split()
                        if len(parts) >= 3:
                            try:
                                deg = float(parts[0])
                                intensity = float(parts[1])
                                degs.append(deg)
                                intensities.append(intensity)
                            except ValueError:
                                continue
                if sim_angle_column is None:
                    sim_angle_column = degs

                try:
                    n_str, f_str = sim_dir.split("_")
                    col_name = f"SIM_design_{n_str}_{f_str}"
                except ValueError:
                    col_name = f"SIM_design_{sim_dir}"

                data[col_name] = intensities

    # === 讀取 Shrinkage 模擬資料 ===
    if os.path.exists(shrinkage_sim_folder):
        shrinkage_sim_structures = sorted(
            [
                d
                for d in os.listdir(shrinkage_sim_folder)
                if os.path.isdir(os.path.join(shrinkage_sim_folder, d))
            ]
        )
        for sim_dir in shrinkage_sim_structures:
            polar_path = os.path.join(shrinkage_sim_folder, sim_dir, polar_tag)
            if os.path.exists(polar_path):
                degs, intensities = [], []
                with open(polar_path, "r") as f:
                    for line in f:
                        parts = line.strip().split()
                        if len(parts) >= 3:
                            try:
                                deg = float(parts[0])
                                intensity = float(parts[1])
                                degs.append(deg)
                                intensities.append(intensity)
                            except ValueError:
                                continue

                if sim_angle_column is None:
                    sim_angle_column = degs

                try:
                    n_str, f_str = sim_dir.split("_")
                    col_name = f"SIM_shrinkage_{n_str}_{f_str}"
                except ValueError:
                    col_name = f"SIM_shrinkage_{sim_dir}"

                data[col_name] = intensities

    # === 加入 SIM_angle ===
    if sim_angle_column is not None:
        data["SIM_angle"] = sim_angle_column
    else:
        data["SIM_angle"] = []

    # === 補齊欄位長度 ===
    if data:
        max_len = max([len(v) for v in data.values()])
        for key in data:
            if len(data[key]) < max_len:
                data[key] = list(data[key]) + [np.nan] * (max_len - len(data[key]))

    # === 寫入工作表 ===
    df_out = pd.DataFrame(data)
    ws = wb.create_sheet(title=angle_str)

    exp_cols = ["EXP_angle"] + sorted(
        [col for col in df_out.columns if col.startswith("EXP")]
    )
    sim_cols = ["SIM_angle"] + sorted(
        [col for col in df_out.columns if col.startswith("SIM")]
    )
    other_cols = [
        col for col in df_out.columns if col not in exp_cols and col not in sim_cols
    ]

    df_out = df_out[exp_cols + sim_cols + other_cols]

    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)


# === 儲存 Excel ===
wb.save(output_path)
print(f"✅ 整理完成，儲存為：{output_path}")


# import os
# import pandas as pd
# import numpy as np
# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# import sys

# sys.stdout.reconfigure(encoding="utf-8")
# # === 設定路徑 ===
# # exp_folder = r"C:\Users\cchih\Desktop\NTHU\MasterThesis\research_log\best_params\MOO_best_PS_OM[0.9, 0.9, 30]\EXP"
# # sim_folder = r"C:\Users\cchih\Desktop\NTHU\MasterThesis\research_log\best_params\MOO_best_PS_OM[0.9, 0.9, 30]\SIM"
# # output_path = r"C:\Users\cchih\Desktop\NTHU\MasterThesis\research_log\best_params\MOO_best_PS_OM[0.9, 0.9, 30]\EXP_SIM.xlsx"
# exp_folder = r"C:\Users\cchih\Desktop\NTHU\MasterThesis\research_log\best_params\MOO_knee_[0.76,0.9,59]\EXP"
# sim_folder = r"C:\Users\cchih\Desktop\NTHU\MasterThesis\research_log\best_params\MOO_knee_[0.76,0.9,59]\SIM"
# output_path = r"C:\Users\cchih\Desktop\NTHU\MasterThesis\research_log\best_params\MOO_knee_[0.76,0.9,59]\EXP_SIM.xlsx"
# # === 仰角範圍 ===
# angles = list(range(10, 90, 10))

# # === 取得所有模擬結構資料夾名稱 ===
# sim_structures = sorted(
#     [d for d in os.listdir(sim_folder) if os.path.isdir(os.path.join(sim_folder, d))]
# )

# # === 建立 Excel 檔案 ===
# wb = Workbook()
# wb.remove(wb.active)  # 移除預設空白工作表

# for angle in angles:
#     angle_str = str(angle)
#     angle_tag = f"ele{angle}"  # EXP檔名關鍵字
#     polar_tag = f"polar-{angle}.txt"  # SIM檔案名稱

#     # === 收集出射角度 ===
#     degree_column = None

#     # === 讀取實驗資料 ===
#     exp_values = {}  # EXP01, EXP02, EXP03
#     angle_exp_files = sorted([f for f in os.listdir(exp_folder) if angle_tag in f])

#     for local_idx, file in enumerate(angle_exp_files):
#         df = pd.read_excel(os.path.join(exp_folder, file))
#         col_name = df.columns[0]  # 第一欄為光角
#         intensity_col = df.columns[1]  # 第二欄為強度
#         if degree_column is None:
#             degree_column = df[col_name].values  # 儲存角度
#         exp_values[f"EXP{local_idx+1:02d}"] = df[intensity_col].values

#     # === 建立工作表資料框 ===
#     data = {"angle": degree_column}
#     data.update(exp_values)

#     # === 讀取模擬資料 ===
#     for sim_dir in sim_structures:
#         polar_path = os.path.join(sim_folder, sim_dir, polar_tag)
#         if os.path.exists(polar_path):
#             degs, intensities = [], []
#             with open(polar_path, "r") as f:
#                 for line in f:
#                     parts = line.strip().split()
#                     if len(parts) >= 3:
#                         try:
#                             deg = float(parts[0])
#                             intensity = float(parts[1])
#                             degs.append(deg)
#                             intensities.append(intensity)
#                         except ValueError:
#                             continue

#             # 拆解資料夾名稱來構建欄位名稱為 SIM_N1.3_F0 格式
#             try:
#                 n_str, f_str = sim_dir.split("_")
#                 col_name = f"SIM_{n_str}_{f_str}"
#             except ValueError:
#                 # 若命名不符預期，則使用原名
#                 col_name = f"SIM_{sim_dir}"

#             data[col_name] = intensities

#     # 找出最長的欄位長度
#     max_len = max([len(v) for v in data.values()])

#     # 將每個欄位補齊到相同長度（使用 NaN）
#     for key in data:
#         if len(data[key]) < max_len:
#             data[key] = list(data[key]) + [np.nan] * (max_len - len(data[key]))

#     # 建立 DataFrame
#     df_out = pd.DataFrame(data)

#     # === 寫入工作表 ===
#     ws = wb.create_sheet(title=angle_str)
#     for r in dataframe_to_rows(df_out, index=False, header=True):
#         ws.append(r)

# # === 儲存 Excel ===
# wb.save(output_path)
# print(f"✅ 整理完成，儲存為：{output_path}")
