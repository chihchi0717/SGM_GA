import pandas as pd
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# === 設定 ===
excel_path = r"C:\Users\cchih\Desktop\NTHU\MasterThesis\GA\SGM_GA\GA_population\best_param0.6_1.2_79\best_param0.6_1.2_79.xlsx"
polar_angles = list(range(10, 90, 10))
txt_folder = r"C:\Users\cchih\Desktop\NTHU\MasterThesis\GA\SGM_GA\GA_population\best_param0.6_1.2_79\N1.3_F2_R44_88"
folder = os.path.dirname(excel_path)
filename = os.path.basename(excel_path)

if filename.startswith("best_param"):
    new_filename = filename.replace("best_param", "update_best_param", 1)
else:
    new_filename = "update_" + filename

output_path = os.path.join(folder, new_filename)
# === 從資料夾名稱提取參數名，作為模擬欄位名稱 ===
folder_name = os.path.basename(txt_folder.rstrip("\\/"))
if "N" in folder_name:
    param_part = folder_name.split("N", 1)[-1]
    sim_label = f"Sim_N{param_part}"
else:
    sim_label = f"Sim_{folder_name}"

# === 檢查與複製檔案 ===
if not os.path.isfile(excel_path):
    raise FileNotFoundError(f"❌ 找不到 Excel 檔案：{excel_path}")
if not os.path.exists(output_path):
    shutil.copyfile(excel_path, output_path)

# === 開啟 Excel 並依序寫入每個仰角工作表 ===
wb = load_workbook(output_path)

for angle in polar_angles:
    sheet_name = str(angle)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        target_col = 1
    else:
        ws = wb[sheet_name]

        # 掃描第1列（標題列），找到最後一欄有標題的 column index
        max_col = ws.max_column
        target_col = 1
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value is not None and str(cell_value).strip() != "":
                target_col = col

        target_col += 1  # 在最後有標題的欄右側保留一欄空白

    # === 讀取 txt 模擬資料 ===
    txt_file = os.path.join(txt_folder, f"polar-{angle}.txt")
    deg_list, intensity_list = [], []

    with open(txt_file, "r") as f:
        for line in f:
            parts = line.strip().split()
            if len(parts) >= 3:
                try:
                    deg = float(parts[0])
                    intensity = float(parts[1])
                    deg_list.append(deg)
                    intensity_list.append(intensity)
                except ValueError:
                    continue

    df = pd.DataFrame({"Sim_Degree": deg_list, sim_label: intensity_list})

    # 寫入模擬資料
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True), start=1
    ):
        for c_idx, value in enumerate(row, start=target_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

# 儲存
wb.save(output_path)
print(f"✅ 模擬資料已寫入：{output_path}，並依據每個工作表中最後一欄自動插入右方。")
